from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.core.files import File
import csv
import io
import os

from .models import Estudiante, Asistencia


# ── Descripción de columnas para mostrar en la plantilla ─────────────────────
COLUMNAS_INFO = [
    {'nombre': 'jornada',            'requerida': False, 'descripcion': 'JM = Jornada Mañana  |  JT = Jornada Tarde  (defecto: JM)'},
    {'nombre': 'tipo',               'requerida': False, 'descripcion': 'CC, TI, PP, OT  (defecto: CC)'},
    {'nombre': 'documento',          'requerida': True,  'descripcion': 'Número de documento. Debe ser único.'},
    {'nombre': 'apellidos',          'requerida': True,  'descripcion': 'Apellidos completos del estudiante'},
    {'nombre': 'nombres',            'requerida': True,  'descripcion': 'Nombres completos del estudiante'},
    {'nombre': 'curso',              'requerida': True,  'descripcion': 'Ej: 10A, 11B, 9C'},
    {'nombre': 'linea',              'requerida': False, 'descripcion': 'AA, ISERC, TPS, COM, ROB, BIO, DIS, OT  (defecto: OT)'},
    {'nombre': 'celular',            'requerida': False, 'descripcion': 'Número de celular del estudiante'},
    {'nombre': 'email',              'requerida': False, 'descripcion': 'Correo electrónico del estudiante'},
    {'nombre': 'acudiente',          'requerida': False, 'descripcion': 'Nombre completo del acudiente'},
    {'nombre': 'parentesco',         'requerida': False, 'descripcion': 'Ej: Madre, Padre, Tío'},
    {'nombre': 'tel_acudiente',      'requerida': False, 'descripcion': 'Teléfono principal del acudiente'},
    {'nombre': 'tel2_acudiente',     'requerida': False, 'descripcion': 'Teléfono secundario del acudiente'},
    {'nombre': 'direccion',          'requerida': False, 'descripcion': 'Dirección de residencia'},
    {'nombre': 'ocupacion_acudiente','requerida': False, 'descripcion': 'Ocupación del acudiente'},
    {'nombre': 'eps',                'requerida': False, 'descripcion': 'Nombre de la EPS'},
    {'nombre': 'observaciones',      'requerida': False, 'descripcion': 'Texto libre con observaciones'},
    {'nombre': 'foto',               'requerida': False, 'descripcion': 'Solo nombre del archivo (ej: Garcia_Juan.jpg). Usar con --carpeta-fotos desde consola.'},
]

COLUMNAS_CSV = [c['nombre'] for c in COLUMNAS_INFO]

FILA_EJEMPLO = {
    'jornada': 'JM', 'tipo': 'TI', 'documento': '1012345678',
    'apellidos': 'García López', 'nombres': 'Juan David', 'curso': '10A',
    'linea': 'TPS', 'celular': '3101234567', 'email': 'juan.garcia@ejemplo.com',
    'acudiente': 'María López', 'parentesco': 'Madre', 'tel_acudiente': '3209876543',
    'tel2_acudiente': '', 'direccion': 'Cra 45 # 12-34', 'ocupacion_acudiente': 'Empleada',
    'eps': 'Sura', 'observaciones': '', 'foto': 'Garcia_Lopez_Juan_David.jpg',
}

VALORES_JORNADA = {v.upper(): k for k, v in Estudiante.JORNADA}
VALORES_JORNADA.update({k: k for k, v in Estudiante.JORNADA})
VALORES_TIPO = {v.upper(): k for k, v in Estudiante.TIPOS_DOCUMENTO}
VALORES_TIPO.update({k: k for k, v in Estudiante.TIPOS_DOCUMENTO})
VALORES_LINEA = {v.upper(): k for k, v in Estudiante.LINEA_MEDIA}
VALORES_LINEA.update({k: k for k, v in Estudiante.LINEA_MEDIA})

CAMPOS_MODELO = [c for c in COLUMNAS_CSV if c != 'foto']


def normalizar(fila):
    if fila.get('jornada'):
        fila['jornada'] = VALORES_JORNADA.get(fila['jornada'].upper(), fila['jornada'])
    if fila.get('tipo'):
        fila['tipo'] = VALORES_TIPO.get(fila['tipo'].upper(), fila['tipo'])
    if fila.get('linea'):
        fila['linea'] = VALORES_LINEA.get(fila['linea'].upper(), fila['linea'])
    return fila


def leer_csv_upload(archivo):
    contenido = archivo.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(contenido))
    return [{k.strip().lower(): (v.strip() if v else '') for k, v in fila.items()} for fila in reader]


def leer_excel_upload(archivo):
    try:
        import openpyxl
    except ImportError:
        raise Exception("Instala openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    encabezados = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {}
        for idx, valor in enumerate(row):
            if idx < len(encabezados) and encabezados[idx]:
                fila[encabezados[idx]] = str(valor).strip() if valor is not None else ''
        if any(v for v in fila.values()):
            filas.append(fila)
    return filas


@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    list_display = ('documento', 'apellidos', 'nombres', 'curso', 'jornada', 'linea')
    search_fields = ('documento', 'apellidos', 'nombres', 'curso')
    list_filter = ('jornada', 'linea', 'tipo')

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('importar/', self.admin_site.admin_view(self.vista_importar), name='importar_estudiantes'),
            path('plantilla/<str:formato>/', self.admin_site.admin_view(self.descargar_plantilla), name='descargar_plantilla'),
        ]
        return custom + urls

    def vista_importar(self, request):
        resultado = None
        if request.method == 'POST':
            archivo = request.FILES.get('archivo')
            actualizar = 'actualizar' in request.POST
            if not archivo:
                messages.error(request, 'Debes seleccionar un archivo.')
            else:
                ext = os.path.splitext(archivo.name)[1].lower()
                try:
                    if ext == '.csv':
                        filas = leer_csv_upload(archivo)
                    elif ext in ('.xlsx', '.xls'):
                        filas = leer_excel_upload(archivo)
                    else:
                        messages.error(request, f'Formato no soportado: {ext}. Usa .csv o .xlsx')
                        filas = []
                    if filas:
                        resultado = self._procesar_filas(filas, actualizar)
                        messages.success(request,
                            f'Importación completada: {resultado["creados"]} creados, '
                            f'{resultado["actualizados"]} actualizados, '
                            f'{resultado["omitidos"]} omitidos.')
                except Exception as e:
                    messages.error(request, f'Error al leer el archivo: {e}')

        contexto = {
            **self.admin_site.each_context(request),
            'title': 'Importar Estudiantes',
            'columnas': COLUMNAS_INFO,
            'resultado': resultado,
            'opts': self.model._meta,
        }
        return render(request, 'admin/estudiantes/importar_estudiantes.html', contexto)

    def _procesar_filas(self, filas, actualizar):
        creados = actualizados = omitidos = 0
        errores = []
        for num, fila in enumerate(filas, start=2):
            fila = normalizar(fila)
            documento = fila.get('documento', '').strip()
            if not documento:
                omitidos += 1
                errores.append(f'Fila {num}: sin documento, se omite.')
                continue
            datos = {c: fila[c] for c in CAMPOS_MODELO if c in fila and fila[c] != ''}
            existe = Estudiante.objects.filter(documento=documento).first()
            if existe and not actualizar:
                omitidos += 1
                continue
            try:
                if existe and actualizar:
                    for campo, valor in datos.items():
                        setattr(existe, campo, valor)
                    existe.save()
                    actualizados += 1
                else:
                    Estudiante.objects.create(**datos)
                    creados += 1
            except Exception as e:
                errores.append(f'Fila {num} ({documento}): {e}')
                omitidos += 1
        return {'creados': creados, 'actualizados': actualizados, 'omitidos': omitidos, 'total': len(filas), 'errores': errores}

    def descargar_plantilla(self, request, formato):
        if formato == 'csv':
            return self._plantilla_csv()
        elif formato == 'xlsx':
            return self._plantilla_excel()
        return HttpResponse('Formato no válido', status=400)

    def _plantilla_csv(self):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.csv"'
        response.write('\ufeff')
        writer = csv.DictWriter(response, fieldnames=COLUMNAS_CSV)
        writer.writeheader()
        writer.writerow(FILA_EJEMPLO)
        return response

    def _plantilla_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('Instala openpyxl: pip install openpyxl', status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'
        for col_idx, nombre in enumerate(COLUMNAS_CSV, start=1):
            celda = ws.cell(row=1, column=col_idx, value=nombre)
            celda.font = Font(bold=True, color='FFFFFF')
            info = next((c for c in COLUMNAS_INFO if c['nombre'] == nombre), None)
            celda.fill = PatternFill('solid', fgColor='C0392B' if (info and info['requerida']) else '2563EB')
            celda.alignment = Alignment(horizontal='center')
        for col_idx, nombre in enumerate(COLUMNAS_CSV, start=1):
            ws.cell(row=2, column=col_idx, value=FILA_EJEMPLO.get(nombre, ''))
        ws2 = wb.create_sheet('Referencia')
        ws2.append(['Columna', 'Obligatoria', 'Descripción / Valores válidos'])
        for h in ['A1','B1','C1']: ws2[h].font = Font(bold=True)
        for info in COLUMNAS_INFO:
            ws2.append([info['nombre'], 'Sí' if info['requerida'] else 'No', info['descripcion']])
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 14
        ws2.column_dimensions['C'].width = 65
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.xlsx"'
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['importar_url'] = 'importar/'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Asistencia)
class AsistenciaAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'fecha', 'hora', 'tipo')
    list_filter = ('tipo', 'fecha')
    search_fields = ('estudiante__documento', 'estudiante__apellidos', 'estudiante__nombres')
