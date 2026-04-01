"""
Comando para importar estudiantes masivamente desde CSV o Excel.

Uso:
    python manage.py importar_estudiantes archivo.csv
    python manage.py importar_estudiantes archivo.xlsx
    python manage.py importar_estudiantes archivo.csv --carpeta-fotos /ruta/a/fotos
    python manage.py importar_estudiantes archivo.csv --actualizar   (actualiza si ya existe el documento)
"""

import os
import csv
from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from estudiantes.models import Estudiante


COLUMNAS_REQUERIDAS = {'documento', 'apellidos', 'nombres', 'curso'}

COLUMNAS_VALIDAS = [
    'jornada', 'tipo', 'documento', 'apellidos', 'nombres', 'curso',
    'linea', 'celular', 'email', 'acudiente', 'parentesco',
    'tel_acudiente', 'tel2_acudiente', 'direccion',
    'ocupacion_acudiente', 'eps', 'observaciones', 'foto',
]

VALORES_JORNADA = {v.upper(): k for k, v in Estudiante.JORNADA}
VALORES_JORNADA.update({k: k for k, v in Estudiante.JORNADA})

VALORES_TIPO = {v.upper(): k for k, v in Estudiante.TIPOS_DOCUMENTO}
VALORES_TIPO.update({k: k for k, v in Estudiante.TIPOS_DOCUMENTO})

VALORES_LINEA = {v.upper(): k for k, v in Estudiante.LINEA_MEDIA}
VALORES_LINEA.update({k: k for k, v in Estudiante.LINEA_MEDIA})


def leer_csv(ruta):
    """Lee un archivo CSV y retorna lista de dicts."""
    filas = []
    with open(ruta, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for fila in reader:
            filas.append({k.strip().lower(): v.strip() for k, v in fila.items()})
    return filas


def leer_excel(ruta):
    """Lee un archivo Excel (.xlsx o .xls) y retorna lista de dicts."""
    try:
        import openpyxl
    except ImportError:
        raise CommandError(
            "Para leer archivos Excel instala openpyxl:\n"
            "    pip install openpyxl"
        )
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = []
    encabezados = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {}
        for idx, valor in enumerate(row):
            if idx < len(encabezados) and encabezados[idx]:
                fila[encabezados[idx]] = str(valor).strip() if valor is not None else ''
        # Ignorar filas completamente vacías
        if any(v for v in fila.values()):
            filas.append(fila)
    return filas


def normalizar_fila(fila):
    """Normaliza los valores de choices al código interno."""
    if 'jornada' in fila and fila['jornada']:
        fila['jornada'] = VALORES_JORNADA.get(fila['jornada'].upper(), fila['jornada'])
    if 'tipo' in fila and fila['tipo']:
        fila['tipo'] = VALORES_TIPO.get(fila['tipo'].upper(), fila['tipo'])
    if 'linea' in fila and fila['linea']:
        fila['linea'] = VALORES_LINEA.get(fila['linea'].upper(), fila['linea'])
    return fila


class Command(BaseCommand):
    help = 'Importa estudiantes masivamente desde un archivo CSV o Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta al archivo CSV o Excel con los datos de los estudiantes',
        )
        parser.add_argument(
            '--carpeta-fotos',
            type=str,
            default=None,
            help='Ruta a la carpeta que contiene las fotos de los estudiantes',
        )
        parser.add_argument(
            '--actualizar',
            action='store_true',
            default=False,
            help='Si el documento ya existe, actualiza el registro en lugar de omitirlo',
        )

    def handle(self, *args, **options):
        ruta_archivo = options['archivo']
        carpeta_fotos = options['carpeta_fotos']
        actualizar = options['actualizar']

        # --- Verificar que el archivo existe ---
        if not os.path.isfile(ruta_archivo):
            raise CommandError(f"No se encontró el archivo: {ruta_archivo}")

        # --- Leer el archivo según extensión ---
        ext = os.path.splitext(ruta_archivo)[1].lower()
        self.stdout.write(f"📂 Leyendo archivo: {ruta_archivo}")

        if ext == '.csv':
            filas = leer_csv(ruta_archivo)
        elif ext in ('.xlsx', '.xls'):
            filas = leer_excel(ruta_archivo)
        else:
            raise CommandError(f"Formato no soportado: {ext}. Use .csv o .xlsx")

        if not filas:
            raise CommandError("El archivo está vacío o no tiene datos.")

        # --- Validar columnas requeridas ---
        columnas_archivo = set(filas[0].keys())
        faltantes = COLUMNAS_REQUERIDAS - columnas_archivo
        if faltantes:
            raise CommandError(
                f"Faltan columnas requeridas: {', '.join(sorted(faltantes))}\n"
                f"Columnas encontradas: {', '.join(sorted(columnas_archivo))}"
            )

        # --- Procesar filas ---
        total = len(filas)
        creados = 0
        actualizados = 0
        omitidos = 0
        errores = []

        self.stdout.write(f"🔄 Procesando {total} registros...\n")

        for num, fila in enumerate(filas, start=2):  # start=2 porque fila 1 es encabezado
            fila = normalizar_fila(fila)
            documento = fila.get('documento', '').strip()

            if not documento:
                errores.append(f"Fila {num}: sin documento, se omite.")
                omitidos += 1
                continue

            # Campos válidos para el modelo
            datos = {
                col: fila[col]
                for col in COLUMNAS_VALIDAS
                if col in fila and col != 'foto' and fila[col] != ''
            }

            # Verificar si ya existe
            existe = Estudiante.objects.filter(documento=documento).first()

            if existe and not actualizar:
                omitidos += 1
                continue

            try:
                if existe and actualizar:
                    for campo, valor in datos.items():
                        setattr(existe, campo, valor)
                    estudiante = existe
                    es_nuevo = False
                else:
                    estudiante = Estudiante(**datos)
                    es_nuevo = True

                # --- Adjuntar foto si se indica ---
                nombre_foto = fila.get('foto', '').strip()
                if nombre_foto and carpeta_fotos:
                    ruta_foto = os.path.join(carpeta_fotos, nombre_foto)
                    if os.path.isfile(ruta_foto):
                        with open(ruta_foto, 'rb') as img:
                            estudiante.foto.save(
                                nombre_foto,
                                File(img),
                                save=False
                            )
                    else:
                        errores.append(f"Fila {num} ({documento}): foto no encontrada → {ruta_foto}")

                estudiante.save()

                if es_nuevo:
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append(f"Fila {num} ({documento}): {e}")
                omitidos += 1

        # --- Resumen final ---
        self.stdout.write(self.style.SUCCESS(f"\n✅ Importación completada"))
        self.stdout.write(f"   Creados:      {creados}")
        self.stdout.write(f"   Actualizados: {actualizados}")
        self.stdout.write(f"   Omitidos:     {omitidos}")
        self.stdout.write(f"   Total filas:  {total}")

        if errores:
            self.stdout.write(self.style.WARNING(f"\n⚠️  Advertencias ({len(errores)}):"))
            for e in errores:
                self.stdout.write(f"   - {e}")
